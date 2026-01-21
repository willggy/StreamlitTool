import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import zipfile
import re

# --- 1. 页面配置与样式 ---
st.set_page_config(page_title="智能分表工具", layout="wide")

st.markdown("""
    <style>
    /* 隐藏右上角的菜单按钮和 GitHub 部署者信息 */
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    .block-container { max-width: 900px !important; margin: 0 auto !important; padding-top: 1.5rem !important; }
    .stApp { background-color: #ffffff; } 
    label[data-testid="stWidgetLabel"] { display: none !important; }

    /* 统一组件高度与底色 */
    div[data-testid="stFileUploader"] section, div[data-testid="stMultiSelect"] > div,
    div[data-testid="stTextInput"] div[data-baseweb="input"], div[data-testid="stMetric"] {
        height: 40px !important; min-height: 40px !important;
        background-color: #ffffff !important; border: 1px solid #d1ccc0 !important; border-radius: 4px !important;
    }

    /* 紧凑间距 */
    [data-testid="stVerticalBlock"] > div { margin-bottom: 15px !important; }

    /* 上传框样式 */
    div[data-testid="stFileUploader"] section { padding: 0px 15px !important; justify-content: flex-start !important; display: flex !important; align-items: center !important; }
    div[data-testid="stFileUploader"] section > div { display: none; } 
    div[data-testid="stFileUploader"] section::after { content: "📎 点击或拖拽上传 Excel 文件"; color: #a39e93; font-size: 14px; margin-left: 5px; }

    /* 指标卡样式 */
    div[data-testid="stMetric"] { padding: 0px 15px !important; display: flex !important; align-items: center !important; justify-content: space-between !important; }
    div[data-testid="stMetricLabel"] { color: #a39e93 !important; font-size: 13px !important; margin: 0 !important; }
    div[data-testid="stMetricValue"] { color: #5a7d9a !important; font-size: 16px !important; padding: 0 !important; }

    /* 按钮样式 */
    .stButton button, .stDownloadButton button { height: 40px !important; border-radius: 4px !important; border: none !important; color: white !important; font-weight: 500 !important; }
    div.stButton > button[kind="primary"] { background-color: #8da4b1 !important; }
    .stDownloadButton button { background-color: #a7ad9b !important; }

    /* 弹出提示居中 */
    div[data-testid="stToast"] { 
        position: fixed !important; top: 50% !important; left: 50% !important; 
        transform: translate(-50%, -50%) !important; width: 320px !important; 
        background-color: #ffffff !important; border: 2px solid #a7ad9b !important; 
        box-shadow: 0 10px 25px rgba(0,0,0,0.1) !important; z-index: 10000 !important; 
    }
    div[data-testid="stHorizontalBlock"] { align-items: center !important; }
    </style>
""", unsafe_allow_html=True)

# --- 2. 核心工具函数 ---
def make_clean_name(prefix, suffix, group_name, sheet_name=""):
    group_part = "-".join(str(v) for v in group_name if pd.notna(v)) if isinstance(group_name, tuple) else str(group_name)
    parts = [p.strip() for p in [prefix, group_part, suffix, sheet_name] if p.strip()]
    name = "-".join(parts)
    return re.sub(r'[\\/*?:[\]]', '_', name)[:31].strip('_- ') or "结果"

def copy_format_and_write(new_ws, orig_ws, group_df):
    for col_letter, dim in orig_ws.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = dim.width
    for row_num in range(1, len(group_df) + 2):
        if row_num in orig_ws.row_dimensions:
            new_ws.row_dimensions[row_num].height = orig_ws.row_dimensions[row_num].height
    for r_idx, row in enumerate(dataframe_to_rows(group_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            new_ws.cell(row=r_idx, column=c_idx, value=value)
    
    max_copy = min(orig_ws.max_row, 100)
    for orig_row in orig_ws.iter_rows(min_row=1, max_row=max_copy):
        for orig_cell in orig_row:
            if orig_cell.number_format != 'General':
                new_ws.cell(row=orig_cell.row, column=orig_cell.column).number_format = orig_cell.number_format
    
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    even_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    for cell in new_ws[1]:
        cell.font, cell.fill, cell.alignment = Font(bold=True), header_fill, Alignment(horizontal="center", vertical="center")
    for row_idx in range(2, len(group_df) + 2):
        fill = even_fill if (row_idx % 2 == 0) else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for cell in new_ws[row_idx]:
            cell.fill, cell.alignment = fill, Alignment(vertical="center")

# --- 3. 界面逻辑 ---
if "res" not in st.session_state: st.session_state.res = None
if "show_success" not in st.session_state: st.session_state.show_success = False

if st.session_state.show_success:
    st.toast("✅ 分表完成！")
    st.session_state.show_success = False

st.markdown("<h2 style='text-align: center; color: #5d5d5d;'>📊 智能分表美化工具</h2>", unsafe_allow_html=True)

uploaded_file = st.file_uploader("upload", type=["xlsx"])
sheet_data = {}
if uploaded_file:
    try:
        wb = load_workbook(uploaded_file, data_only=False)
        for s_name in wb.sheetnames:
            ws = wb[s_name]
            data = list(ws.values)
            df = pd.DataFrame(data[1:], columns=data[0]) if data else pd.DataFrame()
            sheet_data[s_name] = {"df": df, "ws": ws}
    except: st.error("读取失败")

if sheet_data:
    r2c1, r2c2, r2c3 = st.columns([1.5, 1.5, 1])
    
    selected_sheets = r2c1.multiselect("S", options=list(sheet_data.keys()), default=list(sheet_data.keys()))
    
    # 获取选中 Sheet 的共同标题
    common_columns = []
    if selected_sheets:
        list_of_column_sets = [set(sheet_data[s]["df"].columns.tolist()) for s in selected_sheets if not sheet_data[s]["df"].empty]
        if list_of_column_sets:
            intersect_cols = set.intersection(*list_of_column_sets)
            first_sheet_cols = sheet_data[selected_sheets[0]]["df"].columns.tolist()
            common_columns = [col for col in first_sheet_cols if col in intersect_cols]

    group_columns = r2c2.multiselect("C", options=common_columns, placeholder="选择共同关键字列")
    
    # 【核心修改点】计算所有选中 Sheet 的分组并集，确保“按照最大的”计算数量
    all_groups_list = []
    if group_columns and selected_sheets:
        unique_groups = set()
        for s in selected_sheets:
            df_s = sheet_data[s]["df"]
            if not df_s.empty:
                # 提取当前 Sheet 中的分组情况
                current_groups = df_s[group_columns].dropna().drop_duplicates().values
                for g in current_groups:
                    # 转化为元组以便存入 set
                    unique_groups.add(tuple(g) if len(g) > 1 else g[0])
        all_groups_list = sorted(list(unique_groups)) # 转换为列表用于后续迭代

    r2c3.metric("预计数量", f"{len(all_groups_list)}")

    r3c1, r3c2, r3c3 = st.columns([1.2, 1.4, 1.4])
    output_mode = r3c1.radio("M", ["单文件 (多Sheet拆分)", "多文件 (跨Sheet汇总)"], horizontal=True)
    prefix = r3c2.text_input("P", placeholder="前缀")
    suffix = r3c3.text_input("S", placeholder="后缀")

    st.markdown("<br>", unsafe_allow_html=True)
    r4c1, r4c2 = st.columns([1, 1])

    if r4c1.button("⚙️ 开始分表", type="primary", use_container_width=True, disabled=not (group_columns and all_groups_list)):
        with st.spinner("处理中..."):
            if "单文件" in output_mode:
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # 单文件模式下，按选中的 Sheet 逐一处理
                    for s_name in selected_sheets:
                        item = sheet_data[s_name]
                        if item["df"].empty: continue
                        grouped = item["df"].groupby(group_columns, sort=False)
                        for name, group in grouped:
                            s_out = make_clean_name(prefix, suffix, name, s_name)
                            new_ws = writer.book.create_sheet(s_out)
                            copy_format_and_write(new_ws, item["ws"], group)
                output.seek(0)
                st.session_state.res = {"data": output, "name": "分表结果.xlsx"}
            else:
                # 多文件模式：按照最大的并集分组，跨 Sheet 汇总
                zip_buf = BytesIO()
                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zipf:
                    for group_val in all_groups_list:
                        file_name = f"{make_clean_name(prefix, suffix, group_val)}.xlsx"
                        excel_out = BytesIO()
                        with pd.ExcelWriter(excel_out, engine='openpyxl') as writer:
                            has_data = False
                            for s_name in selected_sheets:
                                item = sheet_data[s_name]
                                df_s = item["df"]
                                if df_s.empty: continue
                                
                                # 构建过滤条件
                                vals = group_val if isinstance(group_val, tuple) else [group_val]
                                mask = (df_s[group_columns].astype(str) == [str(v) for v in vals]).all(axis=1)
                                sub_df = df_s[mask]
                                
                                if not sub_df.empty:
                                    new_ws = writer.book.create_sheet(title=s_name)
                                    copy_format_and_write(new_ws, item["ws"], sub_df)
                                    has_data = True
                            
                            if "Sheet" in writer.book.sheetnames: del writer.book["Sheet"]
                        
                        if has_data:
                            zipf.writestr(file_name, excel_out.getvalue())
                zip_buf.seek(0)
                st.session_state.res = {"data": zip_buf, "name": "汇总分表结果.zip"}
            
            st.session_state.show_success = True
            st.rerun()

    if st.session_state.res:

        r4c2.download_button(label="💾 下载结果", data=st.session_state.res["data"], file_name=st.session_state.res["name"], use_container_width=True)

