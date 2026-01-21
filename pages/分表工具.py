import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import zipfile
import re

# --- 1. 页面配置与莫兰迪风格样式 ---
st.set_page_config(page_title="分表工具", layout="wide")

st.markdown("""
    <style>
    /* 隐藏右上角的菜单按钮和 GitHub 部署者信息 */
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    .block-container { max-width: 900px !important; margin: 0 auto !important; padding-top: 2rem !important; }
    .stApp { background-color: #ffffff; } 
    label[data-testid="stWidgetLabel"] { display: none !important; }

    /* 统一组件高度与底色 */
    div[data-testid="stFileUploader"] section,
    div[data-testid="stMultiSelect"] > div,
    div[data-testid="stTextInput"] div[data-baseweb="input"],
    div[data-testid="stMetric"] {
        height: 42px !important; min-height: 42px !important;
        background-color: #ffffff !important; border: 1px solid #d1ccc0 !important; border-radius: 4px !important;
    }

    /* 加宽行间距 */
    [data-testid="stVerticalBlock"] > div { margin-bottom: 25px !important; }

    /* 上传框左对齐 */
    div[data-testid="stFileUploader"] section { padding: 0px 15px !important; justify-content: flex-start !important; display: flex !important; align-items: center !important; }
    div[data-testid="stFileUploader"] section > div { display: none; } 
    div[data-testid="stFileUploader"] section::after { content: "📎 点击或拖拽上传 Excel 文件"; color: #a39e93; font-size: 14px; margin-left: 5px; }

    /* 指标卡同步 */
    div[data-testid="stMetric"] { padding: 0px 15px !important; display: flex !important; align-items: center !important; justify-content: space-between !important; }
    div[data-testid="stMetricLabel"] { color: #a39e93 !important; font-size: 13px !important; margin: 0 !important; }
    div[data-testid="stMetricValue"] { color: #5a7d9a !important; font-size: 16px !important; padding: 0 !important; }

    /* 按钮样式 */
    .stButton button, .stDownloadButton button { height: 42px !important; border-radius: 4px !important; border: none !important; color: white !important; font-weight: 500 !important; }
    div.stButton > button[kind="primary"] { background-color: #8da4b1 !important; }
    .stDownloadButton button { background-color: #a7ad9b !important; }

/* 8. 提示框居中且自带呼吸感 */
div[data-testid="stToast"] { 
    position: fixed !important; 
    top: 50% !important; 
    left: 50% !important; 
    transform: translate(-50%, -50%) !important; 
    width: 320px !important; 
    background-color: #ffffff !important; 
    border: 2px solid #a7ad9b !important; 
    box-shadow: 0 10px 25px rgba(0,0,0,0.15) !important;
    z-index: 10000 !important;
    /* 动画效果：4秒内完成显示和自动隐藏的视觉过渡 */
    animation: toast-fade 2s forwards;
}

@keyframes toast-fade {
    0% { opacity: 0; }
    10% { opacity: 1; }
    90% { opacity: 1; }
    100% { opacity: 0; display: none; }
}
    
    div[data-testid="stHorizontalBlock"] { align-items: center !important; }
    </style>
""", unsafe_allow_html=True)

# --- 2. 核心逻辑函数 ---
def make_name(prefix, suffix, group_name, sheet_name=""):
    if isinstance(group_name, tuple):
        # 处理多列分组的情况，用“-”连接内容
        group_part = "-".join(str(v) for v in group_name if pd.notna(v))
    else:
        group_part = str(group_name)
    
    # 【修改处】移除了列表末尾的 sheet_name
    parts = [p.strip() for p in [prefix, group_part, suffix] if p.strip()]
    
    name = "-".join(parts)
    # 替换 Windows 系统文件名不允许的非法字符，并限制长度（Excel Sheet名上限为31字符）
    name = re.sub(r'[\\/*?:[\]]', '_', name)[:31].strip('_- ')
    
    return name or "结果"  # 如果名字为空，默认返回“结果”

def copy_format_and_write(new_ws, orig_ws, group_df):
    for col_letter, dim in orig_ws.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = dim.width
    for row_num in range(1, len(group_df) + 2):
        if row_num in orig_ws.row_dimensions:
            new_ws.row_dimensions[row_num].height = orig_ws.row_dimensions[row_num].height
    for r_idx, row in enumerate(dataframe_to_rows(group_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            new_ws.cell(row=r_idx, column=c_idx, value=value)
    max_copy = min(orig_ws.max_row, 100)
    for orig_row in orig_ws.iter_rows(min_row=1, max_row=max_copy):
        for orig_cell in orig_row:
            if orig_cell.number_format != 'General':
                new_cell = new_ws.cell(row=orig_cell.row, column=orig_cell.column)
                new_cell.number_format = orig_cell.number_format
    
    # Excel 美化
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    even_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    odd_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for cell in new_ws[1]:
        cell.font, cell.fill, cell.alignment = header_font, header_fill, Alignment(horizontal="center", vertical="center")
    total_rows = len(group_df) + 1
    for row_idx in range(2, total_rows + 1):
        fill = even_fill if (row_idx % 2 == 0) else odd_fill
        for cell in new_ws[row_idx]:
            cell.fill, cell.alignment = fill, Alignment(vertical="center")

# --- 3. 界面逻辑 ---
# 初始化 Session State
if "res" not in st.session_state: st.session_state.res = None
if "show_success" not in st.session_state: st.session_state.show_success = False

# 检查是否需要显示居中弹出框
if st.session_state.show_success:
    st.toast("✅ 分表处理完成！")
    st.session_state.show_success = False # 显示完后重置

st.markdown("<h2 style='text-align: center; color: #5d5d5d;'>📊 Excel 分表工具</h2>", unsafe_allow_html=True)

MAX_FILE_SIZE = 50 * 1024 * 1024
r1c1, r1c2 = st.columns([3, 1])
uploaded_file = r1c1.file_uploader("upload", type=["xlsx"])

sheet_data = {}
if uploaded_file:
    if uploaded_file.size > MAX_FILE_SIZE:
        r1c2.error("超过 50MB")
    else:
        try:
            wb = load_workbook(uploaded_file, data_only=False)
            for s_name in wb.sheetnames:
                ws = wb[s_name]
                data = list(ws.values)
                df = pd.DataFrame(data[1:], columns=data[0]) if data else pd.DataFrame()
                sheet_data[s_name] = {"df": df, "ws": ws}
            r1c2.success(f"已读取 {len(sheet_data)} 个 Sheet")
        except:
            r1c2.error("读取失败")

if sheet_data:
    r2c1, r2c2, r2c3 = st.columns([1.5, 1.5, 1])
    selected_sheets = r2c1.multiselect("S", options=list(sheet_data.keys()), default=list(sheet_data.keys())[:1])
    
    if selected_sheets:
        ref_df = sheet_data[selected_sheets[0]]["df"]
        group_columns = r2c2.multiselect("C", options=ref_df.columns.tolist())
        n_groups = ref_df[group_columns].dropna().drop_duplicates().shape[0] if group_columns else 0
        r2c3.metric("预计数量", f"{n_groups} 个")

        r3c1, r3c2, r3c3 = st.columns([1.2, 1.4, 1.4])
        output_mode = r3c1.radio("M", ["单文件 (多Sheet)", "多文件 (ZIP)"], horizontal=True)
        prefix = r3c2.text_input("P", placeholder="前缀 (可选)")
        suffix = r3c3.text_input("S", placeholder="后缀 (可选)")

        st.markdown("<br>", unsafe_allow_html=True)
        r4c1, r4c2 = st.columns([1, 1])

        if r4c1.button("⚙️ 开始分表", type="primary", use_container_width=True, disabled=not group_columns):
            with st.spinner("处理中..."):
                count = 0
                if "单文件" in output_mode:
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        for s_name in selected_sheets:
                            item = sheet_data[s_name]
                            grouped = item["df"].groupby(group_columns, sort=False)
                            for name, group in grouped:
                                s_out = make_name(prefix, suffix, name, s_name)
                                new_ws = writer.book.create_sheet(s_out)
                                copy_format_and_write(new_ws, item["ws"], group)
                                count += 1
                    output.seek(0)
                    st.session_state.res = {"data": output, "name": "分表结果.xlsx"}
                else:
                    zip_buf = BytesIO()
                    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zipf:
                        for s_name in selected_sheets:
                            item = sheet_data[s_name]
                            grouped = item["df"].groupby(group_columns, sort=False)
                            for name, group in grouped:
                                f_base = make_name(prefix, suffix, name, s_name)
                                buf = BytesIO()
                                with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                                    new_ws = writer.book.create_sheet("Sheet1")
                                    copy_format_and_write(new_ws, item["ws"], group)
                                buf.seek(0)
                                zipf.writestr(f"{f_base}.xlsx", buf.getvalue())
                                count += 1
                    zip_buf.seek(0)
                    st.session_state.res = {"data": zip_buf, "name": "分表结果.zip"}
                
                # 关键：先标记成功，再执行 rerun
                st.session_state.show_success = True
                st.rerun()

        if st.session_state.res:

            r4c2.download_button(label="💾 下载分表结果", data=st.session_state.res["data"], file_name=st.session_state.res["name"], use_container_width=True)
